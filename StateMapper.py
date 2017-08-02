from openpyxl import Workbook, load_workbook
import sys

wb = load_workbook(filename = sys.argv[1])
regions = wb.worksheets

out_book = Workbook()

def stateMap(region1, r1Name, region2, r2Name):
    ws = out_book.create_sheet()
    ws.title = r1Name + "-" + r2Name
    for state1 in region1:
        for state2 in region2:
            if state1 is not state2:
                row = [state1.value, state2.value]
                ws.append(row)

for region1 in regions:
    region1States = []
    for row in region1.iter_rows():
        region1States.append(row[0])
    for region2 in regions:
        if region1 is not region2:
            region2States = []
            for row in region2.iter_rows():
                region2States.append(row[0])
            stateMap(region1States, region1.title, region2States, region2.title)

out_book.save(sys.argv[2])
